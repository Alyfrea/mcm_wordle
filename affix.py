from openpyxl import load_workbook, Workbook
from collections import Counter


filename = './data_new.xlsx'
wb = load_workbook(filename)
sh = wb['Sheet1']

words = [item[0].value for item in sh.iter_rows(
    min_row=2, max_row=360, min_col=3, max_col=3)]

cnt = {k: Counter() for k in 'abcdefghijklmnopqrstuvwxyz'}

alls = set()

for word in words:
    for prev_, next_ in zip(word, word[1:]):
        cnt[prev_][next_] += 1
        # cnt[next_][prev_] += 1

new_wb = Workbook()
new_sh = new_wb.active
new_sh.append(['词缀', '次数'])
for item in words:
    times = 0
    for prev_, next_ in zip(item, item[1:]):
        times += cnt[prev_][next_] ** 2
    new_sh.append([item, times])
new_wb.save('./词缀.xlsx')
