from openpyxl import load_workbook, Workbook
import numpy as np

filename = './data_new.xlsx'
wb = load_workbook(filename)
sh = wb['Sheet1']

words = [item[0].value for item in sh.iter_rows(
    min_row=2, max_row=360, min_col=3, max_col=3)]

new_wb = Workbook()
new_sh = new_wb.active
new_sh.append(['单词', '重复次数'])
for item in words:
    cnt = np.zeros((26,), dtype=int)
    for c in item:
        if 'a' <= c <= 'z':
            cnt[ord(c) - ord('a')] += 1
    ans = 0
    for i in cnt:
        if i != 0:
            ans += 10 ** i
    new_sh.append([item, ans])
new_wb.save('重复次数.xlsx')
