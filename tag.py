from openpyxl import load_workbook, Workbook
import nltk

filename = './data_new.xlsx'
wb = load_workbook(filename)
sh = wb['Sheet1']

words = [item[0].value for item in sh.iter_rows(
    min_row=2, max_row=360, min_col=3, max_col=3)]

ans = nltk.pos_tag(words)

new_wb = Workbook()
new_sh = new_wb.active

for item in ans:
    new_sh.append(item)

new_wb.save('词性.xlsx')
