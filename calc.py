from openpyxl import load_workbook, Workbook
from scipy.stats import norm
from math import sqrt

wb = load_workbook('question2.xlsx')
sh = wb['Sheet1']
nwb = Workbook()
nsh = nwb.active

nsh.append(['期望预测', '方差预测', 'pdf预测1', '预测2', '预测3', '预测4', '预测5', '预测6', '预测7',
            'cdf预测1', '预测2', '预测3', '预测4', '预测5', '预测6', '预测7'])
for ln, item in enumerate(sh.iter_rows(min_col=2, max_col=5, min_row=2, max_row=360), 2):
    tag1 = item[1].value
    tag2 = item[3].value
    tag3 = item[2].value
    tag4 = item[0].value
    exe = 4.596 - 0.001 * tag1 + 0.348 * tag2 - 0.002 * tag3 - 0.001 * tag4
    exs = 1.027 + 0.003 * tag1 - 0.049 * tag2 + 0.0 * tag3 + 0.001 * tag4
    val1 = norm.pdf([1, 2, 3, 4, 5, 6, 7], loc=exe, scale=sqrt(exs))
    val2 = norm.cdf([1, 2, 3, 4, 5, 6, 7], loc=exe, scale=sqrt(exs))
    for i in range(6, 1, -1):
        val2[i] -= val2[i - 1]
    # val /= val.sum()
    # val *= 100
    nsh.append([exe, exs, *val1, *val2])
nwb.save('./question5_res.xlsx')
