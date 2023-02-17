from openpyxl import load_workbook
import numpy as np


filename = './data_new.xlsx'
wb = load_workbook(filename)
sh = wb['Sheet1']

for ln, item in enumerate(sh.iter_rows(min_row=2, max_row=360, min_col=3, max_col=3), 2):
    item = item[0].value
    cnt = np.zeros((26,), dtype=int)
    for c in item:
        if 'a' <= c <= 'z':
            cnt[ord(c) - ord('a')] += 1
    ans = cnt.max()
    sh.cell(row=ln, column=25, value=ans)
wb.save(filename)